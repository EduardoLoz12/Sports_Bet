"""
Build WC 2026 team stats Excel.

Data sources (all free, no API limit):
  1. martj42/international_results — GitHub CSV, all intl. results since 1872
  2. data/static/team_meta.json   — FIFA rank, WC titles, ELO (static public data)
  3. database/sports_agent.db     — goals_1h/2h, corners, cards from api-sports.io

Output: data/wc2026_team_stats.xlsx
"""
import json, sqlite3, sys
from pathlib import Path
from datetime import date, timedelta

import pandas as pd
import requests

BASE   = Path(__file__).parent.parent
DB     = BASE / "database" / "sports_agent.db"
META   = BASE / "data" / "static" / "team_meta.json"
CACHE  = BASE / ".tmp" / "intl_results.csv"
from datetime import datetime as _dt
OUT    = BASE / "data" / f"wc2026_team_stats_{_dt.now().strftime('%H%M%S')}.xlsx"

# ── Martj42 name → football-data.org name mapping ────────────────────────────
MARTJ42_TO_FD = {
    "United States":             "United States",
    "South Korea":               "South Korea",
    "DR Congo":                  "Congo DR",
    "Ivory Coast":               "Ivory Coast",
    "Bosnia and Herzegovina":    "Bosnia-Herzegovina",
    "Czech Republic":            "Czechia",
    "Curacao":                   "Curaçao",
    "Curaçao":              "Curaçao",
    "Cape Verde":                "Cape Verde Islands",
}
# Reverse: football-data.org → martj42
FD_TO_MARTJ42 = {v: k for k, v in MARTJ42_TO_FD.items()}
# Pass-through teams (same name in both datasets)
PASS_THROUGH = [
    "Argentina","France","Spain","England","Brazil","Portugal","Belgium",
    "Netherlands","Germany","Uruguay","Colombia","Croatia","Morocco","Japan",
    "Mexico","Austria","Turkey","Norway","Switzerland","Sweden","Iran",
    "Australia","Ecuador","Senegal","Algeria","Egypt","Ghana","Tunisia",
    "Saudi Arabia","Iraq","Jordan","New Zealand","Uzbekistan","Canada",
    "Panama","Haiti","Scotland","Paraguay","Qatar","South Africa",
    "Bosnia-Herzegovina",
]
for t in PASS_THROUGH:
    if t not in FD_TO_MARTJ42:
        FD_TO_MARTJ42[t] = t


def fetch_csv() -> pd.DataFrame:
    """Download martj42 CSV (cached locally)."""
    if CACHE.exists() and (date.today().isoformat() in CACHE.stat().st_mtime.__str__()):
        print("Using cached intl_results.csv")
    else:
        CACHE.parent.mkdir(parents=True, exist_ok=True)
        print("Downloading martj42/international_results CSV…")
        url = "https://raw.githubusercontent.com/martj42/international_results/master/results.csv"
        r = requests.get(url, timeout=30, verify=False)
        r.raise_for_status()
        CACHE.write_bytes(r.content)
        print(f"  Downloaded {len(r.content)//1024} KB")

    df = pd.read_csv(CACHE, parse_dates=["date"])
    return df


def get_wc_teams() -> list[str]:
    """Get WC 2026 teams from DB."""
    conn = sqlite3.connect(DB)
    rows = conn.execute("""
        SELECT DISTINCT home_team FROM matches WHERE home_team IS NOT NULL
        UNION
        SELECT DISTINCT away_team FROM matches WHERE away_team IS NOT NULL
    """).fetchall()
    conn.close()
    return sorted(r[0] for r in rows)


def get_db_stats() -> dict:
    """Get stats from api-sports.io cached in DB."""
    conn = sqlite3.connect(DB)
    # league_id added later — handle missing column gracefully
    cols = [r[1] for r in conn.execute("PRAGMA table_info(team_stats)").fetchall()]
    league_col = "league_id" if "league_id" in cols else "0"
    rows = conn.execute(f"""
        SELECT team, goals_scored_avg, goals_conceded_avg, goals_1h, goals_2h,
               avg_cards, avg_corners_for, avg_corners_against, {league_col}
        FROM team_stats ORDER BY stat_date DESC
    """).fetchall()
    conn.close()
    seen = set()
    result = {}
    for row in rows:
        team = row[0]
        if team not in seen:
            seen.add(team)
            result[team] = {
                "api_gf_avg":      row[1],
                "api_ga_avg":      row[2],
                "api_goals_1h":    row[3],
                "api_goals_2h":    row[4],
                "api_cards":       row[5] if row[5] and row[5] > 0 else None,
                "api_corners_for": row[6] if row[6] and row[6] > 0 else None,
                "api_corners_ag":  row[7] if row[7] and row[7] > 0 else None,
                "api_league_id":   row[8],
            }
    return result


def calc_team_stats(df: pd.DataFrame, fd_name: str, n_matches: int = 30) -> dict:
    """Calculate stats from martj42 CSV for a given team."""
    import unicodedata
    def norm(s):
        return unicodedata.normalize("NFC", s).casefold()

    m42_name = FD_TO_MARTJ42.get(fd_name, fd_name)

    # Fuzzy fallback: match by normalized string when exact fails
    all_home_teams = df["home_team"].dropna().unique()
    if m42_name not in all_home_teams:
        matches_norm = [t for t in all_home_teams if norm(t) == norm(m42_name)]
        if matches_norm:
            m42_name = matches_norm[0]

    home = df[df["home_team"] == m42_name].copy()
    home["is_home"] = True
    home["gf"] = home["home_score"]
    home["ga"] = home["away_score"]

    away = df[df["away_team"] == m42_name].copy()
    away["is_home"] = False
    away["gf"] = away["away_score"]
    away["ga"] = away["home_score"]

    cols = ["date", "home_team", "away_team", "gf", "ga", "tournament", "neutral", "is_home"]
    matches = pd.concat([home[cols], away[cols]])
    # Drop rows with null scores (unplayed/cancelled matches)
    matches = matches.dropna(subset=["gf", "ga"])
    matches = matches.sort_values("date", ascending=False)

    # Last N matches since 2022
    cutoff = pd.Timestamp("2022-01-01")
    recent = matches[matches["date"] >= cutoff].head(n_matches).copy()

    if recent.empty:
        return {"found": False, "m42_name": m42_name}

    recent["result"] = recent.apply(
        lambda r: "W" if r.gf > r.ga else ("D" if r.gf == r.ga else "L"), axis=1
    )
    recent["clean_sheet"] = recent["ga"] == 0
    recent["big_win"]     = (recent["gf"] - recent["ga"]) >= 3
    recent["big_loss"]    = (recent["ga"] - recent["gf"]) >= 3

    n = len(recent)
    wins   = (recent["result"] == "W").sum()
    draws  = (recent["result"] == "D").sum()
    losses = (recent["result"] == "L").sum()

    recent5  = recent.head(5)["result"].tolist()
    recent10 = recent.head(10)["result"].tolist()
    recent20 = recent.head(20)["result"].tolist()

    def form_str(lst): return "".join(lst)
    def pts(lst): return sum(3 if r == "W" else 1 if r == "D" else 0 for r in lst)
    def pts_pct(lst): return round(pts(lst) / (len(lst) * 3) * 100, 1) if lst else 0

    # Last 5 opponents (for context)
    last5_opponents = []
    for _, row in recent.head(5).iterrows():
        opp = row["away_team"] if row["home_team"] == m42_name else row["home_team"]
        last5_opponents.append(f"{opp}({row['result']})")

    # Goals
    gf_avg = round(recent["gf"].mean(), 2)
    ga_avg = round(recent["ga"].mean(), 2)

    # vs competitive (non-friendly)
    competitive = recent[~recent["tournament"].str.contains("Friendly", case=False, na=False)]
    comp_gf_avg = round(competitive["gf"].mean(), 2) if not competitive.empty else None
    comp_ga_avg = round(competitive["ga"].mean(), 2) if not competitive.empty else None
    comp_w_pct  = round((competitive["result"] == "W").mean() * 100, 1) if not competitive.empty else None

    # Last match date
    last_match = recent.iloc[0]["date"].strftime("%Y-%m-%d")

    return {
        "found":           True,
        "m42_name":        m42_name,
        "total_matches":   n,
        "wins":            int(wins),
        "draws":           int(draws),
        "losses":          int(losses),
        "win_pct":         round(wins / n * 100, 1),
        "draw_pct":        round(draws / n * 100, 1),
        "loss_pct":        round(losses / n * 100, 1),
        "gf_avg":          gf_avg,
        "ga_avg":          ga_avg,
        "goal_diff_avg":   round(gf_avg - ga_avg, 2),
        "comp_gf_avg":     comp_gf_avg,
        "comp_ga_avg":     comp_ga_avg,
        "comp_win_pct":    comp_w_pct,
        "clean_sheet_pct": round(recent["clean_sheet"].mean() * 100, 1),
        "big_win_pct":     round(recent["big_win"].mean() * 100, 1),
        "big_loss_pct":    round(recent["big_loss"].mean() * 100, 1),
        "form5":           form_str(recent5),
        "form10":          form_str(recent10),
        "pts_pct_5":       pts_pct(recent5),
        "pts_pct_10":      pts_pct(recent10),
        "pts_pct_20":      pts_pct(recent20),
        "last5_detail":    " | ".join(last5_opponents),
        "last_match":      last_match,
    }


def main():
    print("Loading data…")
    df       = fetch_csv()
    teams    = get_wc_teams()
    db_stats = get_db_stats()
    meta     = json.loads(META.read_text(encoding="utf-8"))

    print(f"WC 2026 teams from DB: {len(teams)}")
    print(f"Teams with api-sports.io data: {len(db_stats)}")
    print(f"Intl. results rows: {len(df):,}")

    rows = []
    for fd_name in teams:
        s    = calc_team_stats(df, fd_name)
        db   = db_stats.get(fd_name, {})
        m    = meta.get(fd_name, {})

        row = {
            # Identity
            "Team":               fd_name,
            "Confederation":      m.get("confederation", ""),
            "FIFA Rank":          m.get("fifa_rank"),
            "ELO (approx)":       m.get("elo_approx"),
            "WC Titles":          m.get("wc_titles", 0),
            "Continental Titles": m.get("continental_titles", 0),

            # From martj42 (last 30 matches since 2022)
            "Matches (last 30)":  s.get("total_matches"),
            "Win %":              s.get("win_pct"),
            "Draw %":             s.get("draw_pct"),
            "Loss %":             s.get("loss_pct"),
            "GF/game":            s.get("gf_avg"),
            "GA/game":            s.get("ga_avg"),
            "Goal Diff/game":     s.get("goal_diff_avg"),
            "GF/game (compet.)":  s.get("comp_gf_avg"),
            "GA/game (compet.)":  s.get("comp_ga_avg"),
            "Win % (compet.)":    s.get("comp_win_pct"),
            "Clean Sheet %":      s.get("clean_sheet_pct"),
            "Big Win % (3+)":     s.get("big_win_pct"),
            "Big Loss % (3+)":    s.get("big_loss_pct"),
            "Form 5":             s.get("form5"),
            "Form 10":            s.get("form10"),
            "Pts% last 5":        s.get("pts_pct_5"),
            "Pts% last 10":       s.get("pts_pct_10"),
            "Pts% last 20":       s.get("pts_pct_20"),
            "Last 5 (opponent)":  s.get("last5_detail"),
            "Last Match":         s.get("last_match"),

            # From api-sports.io (7 teams only for now)
            "GF/game (api)":      db.get("api_gf_avg"),
            "GA/game (api)":      db.get("api_ga_avg"),
            "Goals 1H/game":      db.get("api_goals_1h"),
            "Goals 2H/game":      db.get("api_goals_2h"),
            "Yellow Cards/game":  db.get("api_cards"),
            "Corners For/game":   db.get("api_corners_for"),
            "Corners Against/game": db.get("api_corners_ag"),
        }
        rows.append(row)
        status = "OK" if s.get("found") else "NOT FOUND"
        print(f"  {fd_name:25s} {status}  matches={s.get('total_matches','?'):>3}  form5={s.get('form5','?')}")

    df_out = pd.DataFrame(rows).sort_values("FIFA Rank")

    # ── Excel output with formatting ──────────────────────────────────────────
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        df_out.to_excel(writer, sheet_name="Team Stats", index=False)

        ws = writer.sheets["Team Stats"]
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Header style
        header_fill = PatternFill("solid", fgColor="1A1A30")
        header_font = Font(color="F0C040", bold=True, size=10)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Row alternating colors
        fill_odd  = PatternFill("solid", fgColor="1A1A24")
        fill_even = PatternFill("solid", fgColor="141420")
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
            for cell in row:
                cell.fill = fill_odd if i % 2 else fill_even
                cell.font = Font(color="E0E0E0", size=9)
                cell.alignment = Alignment(horizontal="center")

        # Column widths
        col_widths = {
            "A": 22, "B": 12, "C": 10, "D": 10, "E": 8, "F": 10,
            "G": 10, "H": 8, "I": 8, "J": 8, "K": 8, "L": 8,
            "M": 10, "N": 10, "O": 10, "P": 10, "Q": 10, "R": 10,
            "S": 10, "T": 8, "U": 10, "V": 10, "W": 10, "X": 10,
            "Y": 30, "Z": 12,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # Freeze header row
        ws.freeze_panes = "B2"

    print(f"\nExcel saved: {OUT}")
    print(f"Rows: {len(df_out)}  |  Columns: {len(df_out.columns)}")


if __name__ == "__main__":
    main()
